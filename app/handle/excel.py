import base64
import json
from openpyxl import load_workbook
import pandas as pd
import os
from fastapi import HTTPException

def check_exist(path):
    exists = os.path.exists(path)
    if exists:
        return True
    else:
        return False

# def TrichXuat_excel(json_front, json_back):
#     data_front = json.loads(json_front)
#     data_back = json.loads(json_back)
    
#     if 'type' in data_front:
#         del data_front['type']
#     if 'type' in data_back:
#         del data_back['type']

#     combined_data = {**data_front, **data_back}

#     df = pd.DataFrame([combined_data])
#     df.to_excel('handle/dir_save/TrichXuatThongTin.xlsx', index=False)

def TrichXuat_excel(json_front, json_back):
    try:
        data_front = json.loads(json_front)
        data_back = json.loads(json_back)
        
        if 'type' in data_front:
            del data_front['type']
        if 'type' in data_back:
            del data_back['type']

        combined_data = {**data_front, **data_back}

        # Tạo DataFrame từ combined_data
        df = pd.DataFrame([combined_data])

        # Sắp xếp lại các cột theo thứ tự mong muốn
        df = df[['name', 'id', 'expire_date', 'birth_day', 
                'gender', 'nationality', 'recent_location', 'origin_location', 
                'issue_place', 'issue_date']]

        # Lưu vào file Excel với index=True
        # df.to_excel('handle/dir_save/TrichXuatThongTin.xlsx', index=True, header=True) 
        df.to_excel('handle/dir_save/TrichXuatThongTin.xlsx', index=False, header=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail='Trích xuất thông tin không thành công, Vui lòng chọn ảnh khác')


def add_row_self(json_front, json_back):
    # workbook = load_workbook(r'D:\learn\VN-cccd-OCR\app\handle\temp_folder\TrichXuatThongTinCoSan.xlsx')
    try:
        workbook = load_workbook(r'handle\temp_folder\TrichXuatThongTinCoSan.xlsx')
        # Select the active sheet
        sheet = workbook.active
        # Chuyển chuỗi JSON thành đối tượng Python
        data_front = json.loads(json_front)
        data_back = json.loads(json_back)
        
        if 'type' in data_front:
            del data_front['type']
        if 'type' in data_back:
            del data_back['type']

        combined_data = {**data_front, **data_back}
        # Tìm hàng trống đầu tiên trong cột 'A'
        next_row = 1
        while sheet[f'A{next_row}'].value is not None:
            next_row += 1

        # Tạo một từ điển ánh xạ các trường JSON với cột tương ứng
        column_mapping = {
            'name': 'A',
            'id': 'B',
            'expire_date': 'C',
            'birth_day': 'D',
            'gender': 'E',
            'nationality': 'F',
            'recent_location': 'G',
            'origin_location': 'H',
            'issue_place': 'I',
            'issue_date': 'J'
        }

        # Duyệt qua dữ liệu và điền vào các cột tương ứng
        for key, value in combined_data.items():
            if key in column_mapping:
                col = column_mapping[key]
                sheet[f'{col}{next_row}'] = value

        # Lưu lại tệp Excel
        workbook.save('handle\dir_save\TrichXuatThongTinCoSan.xlsx')
        # workbook.save(r'D:\learn\VN-cccd-OCR\app\handle\dir_save\TrichXuatThongTinCoSan.xlsx')
        # workbook.save(r'C:\workspace\doanhethongthongminh\VN-cccd-OCR\app\handle\dir_save\TrichXuatThongTinCoSan.xlsx')
    except Exception as e:
        raise HTTPException(status_code=400, detail='Trích xuất thông tin không thành công, Vui lòng chọn ảnh khác')

# done save_excel
def save_excel(excel_data, file_path):
    # Decode base64 thành dữ liệu bytes
    decoded_data_split = excel_data.split(',')
    decoded_data = decoded_data_split[1]
    # Ghi dữ liệu vào tệp Excel
    with open(file_path, "wb") as excel_file:
        excel_file.write(base64.b64decode(decoded_data))
    print("luu thanh cong original excel file")


# base64_str = 'UEsDBBQAAAAIAFq+X1dGWsEMggAAALEAAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE2OTQvCMBBE/0rp3W5V8CAxINSj4Ml7SDc2kGRDdoX8fFPBj9s83jCMuhXKWMQjdzWGxKd+EclHALYLRsND06kZRyUaaVgeQM55ixPZZ8QksBvHA2AVTDPOm/wd7LU65xy8NeIp6au3hZicdJdqMSj4l2vzjoXXvB+2b/lhBb+T+gVQSwMEFAAAAAgAWr5fVxNuJQ7uAAAAKwIAABEAAABkb2NQcm9wcy9jb3JlLnhtbM2SwWrDMAyGX2X4nshOaA8mzaWjpw4GK2zsZmy1NYsTY2skffs5XpsytgfY0dLvT59AjfZSDwGfw+AxkMX4MLmuj1L7DTsTeQkQ9RmdimVK9Kl5HIJTlJ7hBF7pD3VCqDhfg0NSRpGCGVj4hcjaxmipAyoawhVv9IL3n6HLMKMBO3TYUwRRCmDtPNFfpq6BO2CGEQYXvwtoFmKu/onNHWDX5BTtkhrHsRzrnEs7CHh72r/kdQvbR1K9xvQrWkkXjxt2m/xabx8PO9ZWvKoLwYtaHMRarrhcVe+z6w+/u7AbjD3af2x8E2wb+HUX7RdQSwMEFAAAAAgAWr5fV5lcnCMQBgAAnCcAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7Vpbc9o4FH7vr9B4Z/ZtC8Y2gba0E3Npdtu0mYTtTh+FEViNbHlkkYR/v0c2EMuWDe2STbqbPAQs6fvORUfn6Dh58+4uYuiGiJTyeGDZL9vWu7cv3uBXMiQRQTAZp6/wwAqlTF61WmkAwzh9yRMSw9yCiwhLeBTL1lzgWxovI9bqtNvdVoRpbKEYR2RgfV4saEDQVFFab18gtOUfM/gVy1SNZaMBE1dBJrmItPL5bMX82t4+Zc/pOh0ygW4wG1ggf85vp+ROWojhVMLEwGpnP1Zrx9HSSICCyX2UBbpJ9qPTFQgyDTs6nVjOdnz2xO2fjMradDRtGuDj8Xg4tsvSi3AcBOBRu57CnfRsv6RBCbSjadBk2PbarpGmqo1TT9P3fd/rm2icCo1bT9Nrd93TjonGrdB4Db7xT4fDronGq9B062kmJ/2ua6TpFmhCRuPrehIVteVA0yAAWHB21szSA5ZeKfp1lBrZHbvdQVzwWO45iRH+xsUE1mnSGZY0RnKdkAUOADfE0UxQfK9BtorgwpLSXJDWzym1UBoImsiB9UeCIcXcr/31l7vJpDN6nX06zmuUf2mrAaftu5vPk/xz6OSfp5PXTULOcLwsCfH7I1thhyduOxNyOhxnQnzP9vaRpSUyz+/5CutOPGcfVpawXc/P5J6MciO73fZYffZPR24j16nAsyLXlEYkRZ/ILbrkETi1SQ0yEz8InYaYalAcAqQJMZahhvi0xqwR4BN9t74IyN+NiPerb5o9V6FYSdqE+BBGGuKcc+Zz0Wz7B6VG0fZVvNyjl1gVAZcY3zSqNSzF1niVwPGtnDwdExLNlAsGQYaXJCYSqTl+TUgT/iul2v6c00DwlC8k+kqRj2mzI6d0Js3oMxrBRq8bdYdo0jx6/gX5nDUKHJEbHQJnG7NGIYRpu/AerySOmq3CEStCPmIZNhpytRaBtnGphGBaEsbReE7StBH8Waw1kz5gyOzNkXXO1pEOEZJeN0I+Ys6LkBG/HoY4SprtonFYBP2eXsNJweiCy2b9uH6G1TNsLI73R9QXSuQPJqc/6TI0B6OaWQm9hFZqn6qHND6oHjIKBfG5Hj7lengKN5bGvFCugnsB/9HaN8Kr+ILAOX8ufc+l77n0PaHStzcjfWfB04tb3kZuW8T7rjHa1zQuKGNXcs3Ix1SvkynYOZ/A7P1oPp7x7frZJISvmlktIxaQS4GzQSS4/IvK8CrECehkWyUJy1TTZTeKEp5CG27pU/VKldflr7kouDxb5OmvoXQ+LM/5PF/ntM0LM0O3ckvqtpS+tSY4SvSxzHBOHssMO2c8kh22d6AdNfv2XXbkI6UwU5dDuBpCvgNtup3cOjiemJG5CtNSkG/D+enFeBriOdkEuX2YV23n2NHR++fBUbCj7zyWHceI8qIh7qGGmM/DQ4d5e1+YZ5XGUDQUbWysJCxGt2C41/EsFOBkYC2gB4OvUQLyUlVgMVvGAyuQonxMjEXocOeXXF/j0ZLj26ZltW6vKXcZbSJSOcJpmBNnq8reZbHBVR3PVVvysL5qPbQVTs/+Wa3InwwRThYLEkhjlBemSqLzGVO+5ytJxFU4v0UzthKXGLzj5sdxTlO4Ena2DwIyubs5qXplMWem8t8tDAksW4hZEuJNXe3V55ucrnoidvqXd8Fg8v1wyUcP5TvnX/RdQ65+9t3j+m6TO0hMnHnFEQF0RQIjlRwGFhcy5FDukpAGEwHNlMlE8AKCZKYcgJj6C73yDLkpFc6tPjl/RSyDhk5e0iUSFIqwDAUhF3Lj7++TaneM1/osgW2EVDJk1RfKQ4nBPTNyQ9hUJfOu2iYLhdviVM27Gr4mYEvDem6dLSf/217UPbQXPUbzo5ngHrOHc5t6uMJFrP9Y1h75Mt85cNs63gNe5hMsQ6R+wX2KioARq2K+uq9P+SWcO7R78YEgm/zW26T23eAMfNSrWqVkKxE/Swd8H5IGY4xb9DRfjxRiraaxrcbaMQx5gFjzDKFmON+HRZoaM9WLrDmNCm9B1UDlP9vUDWj2DTQckQVeMZm2NqPkTgo83P7vDbDCxI7h7Yu/AVBLAwQUAAAACABavl9X1y+BFpsCAACXBwAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbK1VbU/bMBD+K1Yq9RMiTto0CU4jjTIoaCAGhe0bcpNrY5HEmeOu8O9nu7RDmhM2aZ9857vnuRef7WTLxXNbAEj0UpV1O3UKKZsT122zAiraHvMGamVZcVFRqVSxdttGAM0NqCpdH+OJW1FWO2li9m5FmvCNLFkNtwK1m6qi4vUUSr6dOp6z37hj60LqDTdNGrqGe5APza1QmntgyVkFdct4jQSsps4n7+TK1/7G4ZHBtn0nI13JkvNnrVzmUwfrhKCETGoGqpafMIOy1EQqjR9vnM4hpAa+l/fs56Z2VcuStjDj5TeWy2LqRA7KYUU3pbzj2zm81RMcEjyjkqaJ4FskdJ1pkmnBxDaNUN6s1l26l0JZmQon05pWkLhSpaF1N3tDnfajBGRQy6eSZ1QXayGY9RPAS8MEPOVU2qKf9YOXTMhCYV8t0M/90DXUOQgL7vyjNuk6acmkLehFP5jlFsy8H8MFW7O6r8OXHwRt2w08NSXNbB2++hvwn6fjquk6jJh/GDG/g+ZxOPDHAUGLYjgIo2hE0LUWQo/YCjrtovmiaPwJqdBCbOr1ji32PPJ6hJQl9InRRyTbBwrIEZpr0FjFVmsckrqwDWlXRBy5OFYvzQjbpvMjFA5sg9mFujHZx8Q2lJ2NZaaPMZHohla2iezMceyPcIBxHAe22ub/4wy4pfvonwhs496V2UwDcURmaHE3HIxHnmHCPrm5QL9tWorGanOO7ocDLx6RBfr6sN/cZecRnbcXG6+ZLgEHWnw0zfaJCRBGga+wC83sY4K+a0RgkGEUT8il7bZ1nsbExYGaGN+z3TP33bOuv6xrKtSb0KISVooNH4eBg8TuG9gpkjfmTi+5lLwyYqF+ThDaQdlXnMu9on+hw1+c/gJQSwMEFAAAAAgAWr5fV+DuUEepAgAAFgsAAA0AAAB4bC9zdHlsZXMueG1s3VbbitswEP0V4Q+ok5iauMSBNhAotGVh96GvSizHAllyZTkk+/WdkRznspql7WMTNh7N0ZkzmhnhXfXurMRzI4Rjp1bpvkwa57pPadrvG9Hy/oPphAakNrblDpb2kPadFbzqkdSqdDGb5WnLpU7WKz2029b1bG8G7cpklqTrVW301bNIggO28lawI1dlsuFK7qz0e3kr1Tm4F+jYG2Usc5CKKJM5evrXAM/DCrMc47RSG4vONCiE3924/Qbwjx42SKXuMwPHetVx54TVW1h4jne+gdhov5w7SO1g+Xm++JhcCf4BIjtjK2HvZIJrvVKidkCw8tDg05kuRdA504JRSX4wmvscLoxbJvOtKxPXQOkvYR6dEPPRFQQevZPEaEDme6HUM+76WU/pzyH9U81Cn79W2GKG1byYcObRDGHCAuPfRguxb8Iu/iks6+TRuC8DnEf79a/BOPFkRS1Pfn2qJ30q+pyIDn7eder8WcmDbkU4+x8Lrlf8wmONsfIV1HAK9+AQNmFHYZ3cowca5MtzqscaTeXxxbor/ORleHnK5AfeSXVVZbtBKif1uGpkVQn9pv4Q3vEdXPq7+LC/EjUflHuZwDK52t9FJYe2mHY9YSXGXVf7G87gPJ9uLmhJXYmTqDbj0h523mRggOr48fP7gGz9J45QnIDFEcQoHSoDihNYlM7/dJ4leZ6AUbkto8iS5CxJTmDFkI3/UjpxTgGf+EmLIsvynKroZhPNYEPVLc/xLx6Nyg0ZlA4q/V2t6W7TE/L+HFA9fW9CqJPSk0idlK41IvG6IaMo4t2mdJBBdYGaHdSP6+BMxTlZhl2lcqNuMI0UBYXgLMZnNM+J6uT4jfeHuiVZVhRxBLF4BllGIXgbaYTKAHOgkCzz78GH91F6eU+l1/+E178BUEsDBBQAAAAIAFq+X1eXirscwAAAABMCAAALAAAAX3JlbHMvLnJlbHOdkrluwzAMQH/F0J4wB9AhiDNl8RYE+QFWog/YEgWKRZ2/r9qlcZALGXk9PBLcHmlA7TiktoupGP0QUmla1bgBSLYlj2nOkUKu1CweNYfSQETbY0OwWiw+QC4ZZre9ZBanc6RXiFzXnaU92y9PQW+ArzpMcUJpSEszDvDN0n8y9/MMNUXlSiOVWxp40+X+duBJ0aEiWBaaRcnToh2lfx3H9pDT6a9jIrR6W+j5cWhUCo7cYyWMcWK0/jWCyQ/sfgBQSwMEFAAAAAgAWr5fVxq6G6swAQAAIwIAAA8AAAB4bC93b3JrYm9vay54bWyNUdFKw0AQ/JVwH2BS0YKl6YtFLYgWK32/JJtm6d1t2Nu02q93kxAs+OLT3s4sw8zc8kx8LIiOyZd3IeamEWkXaRrLBryNN9RCUKYm9lZ05UMaWwZbxQZAvEtvs2yeeovBrJaT1pbT64UESkEKCvbAHuEcf/l+TU4YsUCH8p2b4e3AJB4DerxAlZvMJLGh8wsxXiiIdbuSybnczEZiDyxY/oF3vclPW8QBEVt8WDWSm3mmgjVylOFi0Lfq8QR6PG6d0BM6AV5bgWemrsVw6GU0RXoVY+hhmmOJC/5PjVTXWMKays5DkLFHBtcbDLHBNpokWA+5GSwOgXRuqjGcqKurqniBSvCmGv1NpiqoMUD1pjpRcS2o3HLSj0Hn9u5+9qBFdM49KvYeXslWU8bpf1Y/UEsDBBQAAAAIAFq+X1ckHpuirQAAAPgBAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHO1kT0OgzAMha8S5QA1UKlDBUxdWCsuEAXzIxISxa4Kty+FAZA6dGGyni1/78lOn2gUd26gtvMkRmsGymTL7O8ApFu0ii7O4zBPahes4lmGBrzSvWoQkii6QdgzZJ7umaKcPP5DdHXdaXw4/bI48A8wvF3oqUVkKUoVGuRMwmi2NsFS4stMlqKoMhmKKpZwWiDiySBtaVZ9sE9OtOd5Fzf3Ra7N4wmu3wxweHT+AVBLAwQUAAAACABavl9XZZB5khkBAADPAwAAEwAAAFtDb250ZW50X1R5cGVzXS54bWytk01OwzAQha8SZVslLixYoKYbYAtdcAFjTxqr/pNnWtLbM07aSqASFYVNrHjevM+el6zejxGw6J312JQdUXwUAlUHTmIdIniutCE5SfyatiJKtZNbEPfL5YNQwRN4qih7lOvVM7Ryb6l46XkbTfBNmcBiWTyNwsxqShmjNUoS18XB6x+U6kSouXPQYGciLlhQiquEXPkdcOp7O0BKRkOxkYlepWOV6K1AOlrAetriyhlD2xoFOqi945YaYwKpsQMgZ+vRdDFNJp4wjM+72fzBZgrIyk0KETmxBH/HnSPJ3VVkI0hkpq94IbL17PtBTluDvpHN4/0MaTfkgWJY5s/4e8YX/xvO8RHC7r8/sbzWThp/5ovhP15/AVBLAQIUABQAAAAIAFq+X1dGWsEMggAAALEAAAAQAAAAAAAAAAAAAACAAQAAAABkb2NQcm9wcy9hcHAueG1sUEsBAhQAFAAAAAgAWr5fVxNuJQ7uAAAAKwIAABEAAAAAAAAAAAAAAIABsAAAAGRvY1Byb3BzL2NvcmUueG1sUEsBAhQAFAAAAAgAWr5fV5lcnCMQBgAAnCcAABMAAAAAAAAAAAAAAIABzQEAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECFAAUAAAACABavl9X1y+BFpsCAACXBwAAGAAAAAAAAAAAAAAAtoEOCAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQAFAAAAAgAWr5fV+DuUEepAgAAFgsAAA0AAAAAAAAAAAAAAIAB3woAAHhsL3N0eWxlcy54bWxQSwECFAAUAAAACABavl9Xl4q7HMAAAAATAgAACwAAAAAAAAAAAAAAgAGzDQAAX3JlbHMvLnJlbHNQSwECFAAUAAAACABavl9XGrobqzABAAAjAgAADwAAAAAAAAAAAAAAgAGcDgAAeGwvd29ya2Jvb2sueG1sUEsBAhQAFAAAAAgAWr5fVyQem6KtAAAA+AEAABoAAAAAAAAAAAAAAIAB+Q8AAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQAFAAAAAgAWr5fV2WQeZIZAQAAzwMAABMAAAAAAAAAAAAAAIAB3hAAAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAAAkACQA+AgAAKBIAAAAA'
# excel_data = 'data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,UEsDBBQAAAAIAFq+X1dGWsEMggAAALEAAAAQAAAAZG9jUHJvcHMvYXBwLnhtbE2OTQvCMBBE/0rp3W5V8CAxINSj4Ml7SDc2kGRDdoX8fFPBj9s83jCMuhXKWMQjdzWGxKd+EclHALYLRsND06kZRyUaaVgeQM55ixPZZ8QksBvHA2AVTDPOm/wd7LU65xy8NeIp6au3hZicdJdqMSj4l2vzjoXXvB+2b/lhBb+T+gVQSwMEFAAAAAgAWr5fVxNuJQ7uAAAAKwIAABEAAABkb2NQcm9wcy9jb3JlLnhtbM2SwWrDMAyGX2X4nshOaA8mzaWjpw4GK2zsZmy1NYsTY2skffs5XpsytgfY0dLvT59AjfZSDwGfw+AxkMX4MLmuj1L7DTsTeQkQ9RmdimVK9Kl5HIJTlJ7hBF7pD3VCqDhfg0NSRpGCGVj4hcjaxmipAyoawhVv9IL3n6HLMKMBO3TYUwRRCmDtPNFfpq6BO2CGEQYXvwtoFmKu/onNHWDX5BTtkhrHsRzrnEs7CHh72r/kdQvbR1K9xvQrWkkXjxt2m/xabx8PO9ZWvKoLwYtaHMRarrhcVe+z6w+/u7AbjD3af2x8E2wb+HUX7RdQSwMEFAAAAAgAWr5fV5lcnCMQBgAAnCcAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7Vpbc9o4FH7vr9B4Z/ZtC8Y2gba0E3Npdtu0mYTtTh+FEViNbHlkkYR/v0c2EMuWDe2STbqbPAQs6fvORUfn6Dh58+4uYuiGiJTyeGDZL9vWu7cv3uBXMiQRQTAZp6/wwAqlTF61WmkAwzh9yRMSw9yCiwhLeBTL1lzgWxovI9bqtNvdVoRpbKEYR2RgfV4saEDQVFFab18gtOUfM/gVy1SNZaMBE1dBJrmItPL5bMX82t4+Zc/pOh0ygW4wG1ggf85vp+ROWojhVMLEwGpnP1Zrx9HSSICCyX2UBbpJ9qPTFQgyDTs6nVjOdnz2xO2fjMradDRtGuDj8Xg4tsvSi3AcBOBRu57CnfRsv6RBCbSjadBk2PbarpGmqo1TT9P3fd/rm2icCo1bT9Nrd93TjonGrdB4Db7xT4fDronGq9B062kmJ/2ua6TpFmhCRuPrehIVteVA0yAAWHB21szSA5ZeKfp1lBrZHbvdQVzwWO45iRH+xsUE1mnSGZY0RnKdkAUOADfE0UxQfK9BtorgwpLSXJDWzym1UBoImsiB9UeCIcXcr/31l7vJpDN6nX06zmuUf2mrAaftu5vPk/xz6OSfp5PXTULOcLwsCfH7I1thhyduOxNyOhxnQnzP9vaRpSUyz+/5CutOPGcfVpawXc/P5J6MciO73fZYffZPR24j16nAsyLXlEYkRZ/ILbrkETi1SQ0yEz8InYaYalAcAqQJMZahhvi0xqwR4BN9t74IyN+NiPerb5o9V6FYSdqE+BBGGuKcc+Zz0Wz7B6VG0fZVvNyjl1gVAZcY3zSqNSzF1niVwPGtnDwdExLNlAsGQYaXJCYSqTl+TUgT/iul2v6c00DwlC8k+kqRj2mzI6d0Js3oMxrBRq8bdYdo0jx6/gX5nDUKHJEbHQJnG7NGIYRpu/AerySOmq3CEStCPmIZNhpytRaBtnGphGBaEsbReE7StBH8Waw1kz5gyOzNkXXO1pEOEZJeN0I+Ys6LkBG/HoY4SprtonFYBP2eXsNJweiCy2b9uH6G1TNsLI73R9QXSuQPJqc/6TI0B6OaWQm9hFZqn6qHND6oHjIKBfG5Hj7lengKN5bGvFCugnsB/9HaN8Kr+ILAOX8ufc+l77n0PaHStzcjfWfB04tb3kZuW8T7rjHa1zQuKGNXcs3Ix1SvkynYOZ/A7P1oPp7x7frZJISvmlktIxaQS4GzQSS4/IvK8CrECehkWyUJy1TTZTeKEp5CG27pU/VKldflr7kouDxb5OmvoXQ+LM/5PF/ntM0LM0O3ckvqtpS+tSY4SvSxzHBOHssMO2c8kh22d6AdNfv2XXbkI6UwU5dDuBpCvgNtup3cOjiemJG5CtNSkG/D+enFeBriOdkEuX2YV23n2NHR++fBUbCj7zyWHceI8qIh7qGGmM/DQ4d5e1+YZ5XGUDQUbWysJCxGt2C41/EsFOBkYC2gB4OvUQLyUlVgMVvGAyuQonxMjEXocOeXXF/j0ZLj26ZltW6vKXcZbSJSOcJpmBNnq8reZbHBVR3PVVvysL5qPbQVTs/+Wa3InwwRThYLEkhjlBemSqLzGVO+5ytJxFU4v0UzthKXGLzj5sdxTlO4Ena2DwIyubs5qXplMWem8t8tDAksW4hZEuJNXe3V55ucrnoidvqXd8Fg8v1wyUcP5TvnX/RdQ65+9t3j+m6TO0hMnHnFEQF0RQIjlRwGFhcy5FDukpAGEwHNlMlE8AKCZKYcgJj6C73yDLkpFc6tPjl/RSyDhk5e0iUSFIqwDAUhF3Lj7++TaneM1/osgW2EVDJk1RfKQ4nBPTNyQ9hUJfOu2iYLhdviVM27Gr4mYEvDem6dLSf/217UPbQXPUbzo5ngHrOHc5t6uMJFrP9Y1h75Mt85cNs63gNe5hMsQ6R+wX2KioARq2K+uq9P+SWcO7R78YEgm/zW26T23eAMfNSrWqVkKxE/Swd8H5IGY4xb9DRfjxRiraaxrcbaMQx5gFjzDKFmON+HRZoaM9WLrDmNCm9B1UDlP9vUDWj2DTQckQVeMZm2NqPkTgo83P7vDbDCxI7h7Yu/AVBLAwQUAAAACABavl9X1y+BFpsCAACXBwAAGAAAAHhsL3dvcmtzaGVldHMvc2hlZXQxLnhtbK1VbU/bMBD+K1Yq9RMiTto0CU4jjTIoaCAGhe0bcpNrY5HEmeOu8O9nu7RDmhM2aZ9857vnuRef7WTLxXNbAEj0UpV1O3UKKZsT122zAiraHvMGamVZcVFRqVSxdttGAM0NqCpdH+OJW1FWO2li9m5FmvCNLFkNtwK1m6qi4vUUSr6dOp6z37hj60LqDTdNGrqGe5APza1QmntgyVkFdct4jQSsps4n7+TK1/7G4ZHBtn0nI13JkvNnrVzmUwfrhKCETGoGqpafMIOy1EQqjR9vnM4hpAa+l/fs56Z2VcuStjDj5TeWy2LqRA7KYUU3pbzj2zm81RMcEjyjkqaJ4FskdJ1pkmnBxDaNUN6s1l26l0JZmQon05pWkLhSpaF1N3tDnfajBGRQy6eSZ1QXayGY9RPAS8MEPOVU2qKf9YOXTMhCYV8t0M/90DXUOQgL7vyjNuk6acmkLehFP5jlFsy8H8MFW7O6r8OXHwRt2w08NSXNbB2++hvwn6fjquk6jJh/GDG/g+ZxOPDHAUGLYjgIo2hE0LUWQo/YCjrtovmiaPwJqdBCbOr1ji32PPJ6hJQl9InRRyTbBwrIEZpr0FjFVmsckrqwDWlXRBy5OFYvzQjbpvMjFA5sg9mFujHZx8Q2lJ2NZaaPMZHohla2iezMceyPcIBxHAe22ub/4wy4pfvonwhs496V2UwDcURmaHE3HIxHnmHCPrm5QL9tWorGanOO7ocDLx6RBfr6sN/cZecRnbcXG6+ZLgEHWnw0zfaJCRBGga+wC83sY4K+a0RgkGEUT8il7bZ1nsbExYGaGN+z3TP33bOuv6xrKtSb0KISVooNH4eBg8TuG9gpkjfmTi+5lLwyYqF+ThDaQdlXnMu9on+hw1+c/gJQSwMEFAAAAAgAWr5fV+DuUEepAgAAFgsAAA0AAAB4bC9zdHlsZXMueG1s3VbbitswEP0V4Q+ok5iauMSBNhAotGVh96GvSizHAllyZTkk+/WdkRznspql7WMTNh7N0ZkzmhnhXfXurMRzI4Rjp1bpvkwa57pPadrvG9Hy/oPphAakNrblDpb2kPadFbzqkdSqdDGb5WnLpU7WKz2029b1bG8G7cpklqTrVW301bNIggO28lawI1dlsuFK7qz0e3kr1Tm4F+jYG2Usc5CKKJM5evrXAM/DCrMc47RSG4vONCiE3924/Qbwjx42SKXuMwPHetVx54TVW1h4jne+gdhov5w7SO1g+Xm++JhcCf4BIjtjK2HvZIJrvVKidkCw8tDg05kuRdA504JRSX4wmvscLoxbJvOtKxPXQOkvYR6dEPPRFQQevZPEaEDme6HUM+76WU/pzyH9U81Cn79W2GKG1byYcObRDGHCAuPfRguxb8Iu/iks6+TRuC8DnEf79a/BOPFkRS1Pfn2qJ30q+pyIDn7eder8WcmDbkU4+x8Lrlf8wmONsfIV1HAK9+AQNmFHYZ3cowca5MtzqscaTeXxxbor/ORleHnK5AfeSXVVZbtBKif1uGpkVQn9pv4Q3vEdXPq7+LC/EjUflHuZwDK52t9FJYe2mHY9YSXGXVf7G87gPJ9uLmhJXYmTqDbj0h523mRggOr48fP7gGz9J45QnIDFEcQoHSoDihNYlM7/dJ4leZ6AUbkto8iS5CxJTmDFkI3/UjpxTgGf+EmLIsvynKroZhPNYEPVLc/xLx6Nyg0ZlA4q/V2t6W7TE/L+HFA9fW9CqJPSk0idlK41IvG6IaMo4t2mdJBBdYGaHdSP6+BMxTlZhl2lcqNuMI0UBYXgLMZnNM+J6uT4jfeHuiVZVhRxBLF4BllGIXgbaYTKAHOgkCzz78GH91F6eU+l1/+E178BUEsDBBQAAAAIAFq+X1eXirscwAAAABMCAAALAAAAX3JlbHMvLnJlbHOdkrluwzAMQH/F0J4wB9AhiDNl8RYE+QFWog/YEgWKRZ2/r9qlcZALGXk9PBLcHmlA7TiktoupGP0QUmla1bgBSLYlj2nOkUKu1CweNYfSQETbY0OwWiw+QC4ZZre9ZBanc6RXiFzXnaU92y9PQW+ArzpMcUJpSEszDvDN0n8y9/MMNUXlSiOVWxp40+X+duBJ0aEiWBaaRcnToh2lfx3H9pDT6a9jIrR6W+j5cWhUCo7cYyWMcWK0/jWCyQ/sfgBQSwMEFAAAAAgAWr5fVxq6G6swAQAAIwIAAA8AAAB4bC93b3JrYm9vay54bWyNUdFKw0AQ/JVwH2BS0YKl6YtFLYgWK32/JJtm6d1t2Nu02q93kxAs+OLT3s4sw8zc8kx8LIiOyZd3IeamEWkXaRrLBryNN9RCUKYm9lZ05UMaWwZbxQZAvEtvs2yeeovBrJaT1pbT64UESkEKCvbAHuEcf/l+TU4YsUCH8p2b4e3AJB4DerxAlZvMJLGh8wsxXiiIdbuSybnczEZiDyxY/oF3vclPW8QBEVt8WDWSm3mmgjVylOFi0Lfq8QR6PG6d0BM6AV5bgWemrsVw6GU0RXoVY+hhmmOJC/5PjVTXWMKays5DkLFHBtcbDLHBNpokWA+5GSwOgXRuqjGcqKurqniBSvCmGv1NpiqoMUD1pjpRcS2o3HLSj0Hn9u5+9qBFdM49KvYeXslWU8bpf1Y/UEsDBBQAAAAIAFq+X1ckHpuirQAAAPgBAAAaAAAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHO1kT0OgzAMha8S5QA1UKlDBUxdWCsuEAXzIxISxa4Kty+FAZA6dGGyni1/78lOn2gUd26gtvMkRmsGymTL7O8ApFu0ii7O4zBPahes4lmGBrzSvWoQkii6QdgzZJ7umaKcPP5DdHXdaXw4/bI48A8wvF3oqUVkKUoVGuRMwmi2NsFS4stMlqKoMhmKKpZwWiDiySBtaVZ9sE9OtOd5Fzf3Ra7N4wmu3wxweHT+AVBLAwQUAAAACABavl9XZZB5khkBAADPAwAAEwAAAFtDb250ZW50X1R5cGVzXS54bWytk01OwzAQha8SZVslLixYoKYbYAtdcAFjTxqr/pNnWtLbM07aSqASFYVNrHjevM+el6zejxGw6J312JQdUXwUAlUHTmIdIniutCE5SfyatiJKtZNbEPfL5YNQwRN4qih7lOvVM7Ryb6l46XkbTfBNmcBiWTyNwsxqShmjNUoS18XB6x+U6kSouXPQYGciLlhQiquEXPkdcOp7O0BKRkOxkYlepWOV6K1AOlrAetriyhlD2xoFOqi945YaYwKpsQMgZ+vRdDFNJp4wjM+72fzBZgrIyk0KETmxBH/HnSPJ3VVkI0hkpq94IbL17PtBTluDvpHN4/0MaTfkgWJY5s/4e8YX/xvO8RHC7r8/sbzWThp/5ovhP15/AVBLAQIUABQAAAAIAFq+X1dGWsEMggAAALEAAAAQAAAAAAAAAAAAAACAAQAAAABkb2NQcm9wcy9hcHAueG1sUEsBAhQAFAAAAAgAWr5fVxNuJQ7uAAAAKwIAABEAAAAAAAAAAAAAAIABsAAAAGRvY1Byb3BzL2NvcmUueG1sUEsBAhQAFAAAAAgAWr5fV5lcnCMQBgAAnCcAABMAAAAAAAAAAAAAAIABzQEAAHhsL3RoZW1lL3RoZW1lMS54bWxQSwECFAAUAAAACABavl9X1y+BFpsCAACXBwAAGAAAAAAAAAAAAAAAtoEOCAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAhQAFAAAAAgAWr5fV+DuUEepAgAAFgsAAA0AAAAAAAAAAAAAAIAB3woAAHhsL3N0eWxlcy54bWxQSwECFAAUAAAACABavl9Xl4q7HMAAAAATAgAACwAAAAAAAAAAAAAAgAGzDQAAX3JlbHMvLnJlbHNQSwECFAAUAAAACABavl9XGrobqzABAAAjAgAADwAAAAAAAAAAAAAAgAGcDgAAeGwvd29ya2Jvb2sueG1sUEsBAhQAFAAAAAgAWr5fVyQem6KtAAAA+AEAABoAAAAAAAAAAAAAAIAB+Q8AAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzUEsBAhQAFAAAAAgAWr5fV2WQeZIZAQAAzwMAABMAAAAAAAAAAAAAAIAB3hAAAFtDb250ZW50X1R5cGVzXS54bWxQSwUGAAAAAAkACQA+AgAAKBIAAAAA'
# output_path = r'D:\learn\VN-cccd-OCR\app\handle\temp_folder\TrichXuatThongTinCoSan.xlsx'
# save_excel(excel_data, output_path)
